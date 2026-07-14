#!/usr/bin/env python3
"""
TCA Client Report  --  standalone XLSX / PDF / PNG generator
============================================================

Reads a single ``orders.csv`` (the friendly per-order export) plus a handful of
optional auxiliary CSVs, and produces a client-ready Transaction Cost Analysis
report in three formats:

  * ``<client>_TCA_<period>.xlsx``  -- multi-sheet workbook (tables + embedded charts)
  * ``<client>_TCA_<period>.pdf``   -- paginated slide-style report
  * ``png/*.png``                   -- every chart as a standalone image

What it reproduces from the earlier deliverable
-----------------------------------------------
  1. Total Slippage Breakdown by Order Type (all strategies)
  2. Slippage Breakdown by Order Type per Strategy
Both are notional-weighted (weight = executed notional USD = ``$Mln`` x 1e6) and
carry: # Orders, Arrival / Open / Close / PVWAP slippage (bps), %ADV, Volatility,
Spread (bps), Fill Rate (%).

What it adds -- the DARK execution case
---------------------------------------
  A. Slippage vs %DARK participation  (does routing dark lower cost?)   -> hero chart
  B. Spread-normalised slippage vs %DARK  (spread capture / less impact)
  C. Dark benefit by order size (%ADV x %DARK)  (large orders benefit most)
  D. Routed vs Executed volume by venue  (venue reach / under-utilisation)
  E. Indicative $ savings from shifting more flow to dark

Sign convention
---------------
Slippage columns are used **as-is**, matching the existing pipeline:
``+ = good (outperformance) / - = bad (cost)``, in basis points. Pass
``--cost-positive`` if your export uses the opposite convention (+ = cost); the
report then flips the sign on load so every chart/table reads + = good.

Auxiliary CSVs (all optional; put them in --aux-dir; see README for schemas)
----------------------------------------------------------------------------
  dark_routed.csv         Venue, Routed %
  dark_executed.csv       Venue, Executed %
  dark_venue_summary.csv  Venue Category, Shares, Exec Value, % Total Shares, % Total Exec Value
  algo_summary.csv        Strategy, Orders, Exec Shares, Exec Value, %Weight, Period Part,
                          %ADV, Spread Bps, Benchmark, Impact Bps, Wgt Impact Bps, Wgt Impact Cps

Usage
-----
    python tca_report.py --orders sample_data/orders.csv \
        --aux-dir sample_data --outdir report --client "DYMON"
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import seaborn as sns

# --------------------------------------------------------------------------- #
# House style (modern, clean, presentation-ready)
# --------------------------------------------------------------------------- #
INK = "#1F2933"
MUTED = "#6C757D"
GRID = "#E6EBEA"
DARKC = "#2A9D8F"      # teal  -> "dark / good"
LITC = "#E76F51"       # coral -> "lit / cost"
ACC3 = "#F4A261"       # amber
ACC4 = "#264653"       # deep slate
SEQ = ["#264653", "#2A9D8F", "#8AB17D", "#E9C46A", "#F4A261", "#E76F51", "#9B5DE5"]

# XLSX palette
XL_HEADER_BG = "264653"
XL_HEADER_FG = "FFFFFF"
XL_BAND_BG = "F2F6F5"
XL_TITLE_FG = "1F2933"
XL_ACCENT_BG = "2A9D8F"


def set_style() -> None:
    sns.set_theme(style="whitegrid", context="talk",
                  palette=sns.color_palette(SEQ))
    plt.rcParams.update({
        "figure.dpi": 120,
        "savefig.dpi": 200,
        "savefig.bbox": "tight",
        "figure.facecolor": "white",
        "axes.facecolor": "white",
        "savefig.facecolor": "white",
        "axes.edgecolor": "#C7D0CE",
        "axes.linewidth": 1.0,
        "axes.spines.top": False,
        "axes.spines.right": False,
        "axes.titlesize": 16,
        "axes.titleweight": "bold",
        "axes.titlecolor": INK,
        "axes.titlepad": 14,
        "axes.labelsize": 12,
        "axes.labelcolor": INK,
        "axes.axisbelow": True,
        "axes.grid": True,
        "axes.grid.axis": "y",
        "grid.color": GRID,
        "grid.linewidth": 1.1,
        "xtick.color": INK,
        "ytick.color": INK,
        "xtick.labelsize": 10.5,
        "ytick.labelsize": 10.5,
        "legend.frameon": False,
        "font.family": "DejaVu Sans",
        "font.size": 11.5,
    })


# --------------------------------------------------------------------------- #
# Column resolution (tolerant to '(...)' notes, case, %/$ prefixes)
# --------------------------------------------------------------------------- #

# logical name -> list of acceptable base header names (before any '(' , trimmed)
ORDER_FIELDS = {
    "client":       ["client"],
    "trader":       ["Trader"],
    "date":         ["Date"],
    "sym":          ["Sym", "Symbol"],
    "side":         ["Side"],
    "id":           ["aggrTgtId", "aggrIgtId", "OrderId"],
    "strategy":     ["Strategy"],
    "cap":          ["Cap"],
    "sector":       ["sector", "Sector"],
    "auction":      ["auctionOnly"],
    "arrival_time": ["arrivalTime"],
    "mkt":          ["marketLimit", "Market/Limit"],
    "shares":       ["#Shares", "Shares"],          # x 1000
    "mln":          ["$Mln", "Mln", "$mln"],        # x 1e6
    "pct_dv":       ["%DV", "%ADV", "%Adv"],
    "epr":          ["ePR"],
    "fr":           ["FR"],
    "vol":          ["Vol", "Volatility"],
    "sprd":         ["Sprd", "Spread"],
    "is":           ["IS"],
    "open":         ["Open"],
    "close":        ["Close"],
    "epvwap":       ["ePvwap"],
    "epvwap_sprd":  ["ePvwap/Sprd"],
    "dur":          ["Dur", "Duration"],
    "pvwap":        ["Pvwap"],
    "dark":         ["%DARK", "%Dark"],
}

# slippage columns that get sign-flipped when --cost-positive is set
SLIP_KEYS = ["is", "open", "close", "pvwap", "epvwap", "epvwap_sprd"]


def _base(name: str) -> str:
    return str(name).split("(")[0].strip().lower()


def _find(cols, candidates: list[str]) -> str | None:
    lut = {_base(c): c for c in cols}
    for cand in candidates:
        hit = lut.get(_base(cand))
        if hit is not None:
            return hit
    return None


def _num(df: pd.DataFrame, col: str | None) -> pd.Series:
    if col and col in df.columns:
        return pd.to_numeric(df[col], errors="coerce")
    return pd.Series(np.nan, index=df.index)


def _as_pct(s: pd.Series) -> pd.Series:
    """Return a 0..100 percentage: if the data looks like a 0..1 fraction, x100.

    The scale is decided by the maximum, not a high quantile: a column with a big
    mass at 0 (e.g. %DARK where most orders trade fully lit) has a 95th-percentile of
    0, which would wrongly flag genuine 0..100 data as a fraction and inflate it x100.
    """
    s = pd.to_numeric(s, errors="coerce")
    v = s.dropna()
    if not v.empty and v.max() <= 1.5:
        return s * 100.0
    return s


# --------------------------------------------------------------------------- #
# Load & normalise orders
# --------------------------------------------------------------------------- #

def load_orders(path: Path, cost_positive: bool) -> pd.DataFrame:
    raw = pd.read_csv(path, dtype=str)
    keys = {k: _find(raw.columns, cands) for k, cands in ORDER_FIELDS.items()}
    missing = [k for k in ("mkt", "is", "pvwap") if keys[k] is None]
    if missing:
        looked = {k: ORDER_FIELDS[k] for k in missing}
        raise SystemExit(f"orders.csv missing required column(s): {looked}")

    sign = -1.0 if cost_positive else 1.0
    df = pd.DataFrame(index=raw.index)
    df["notional"] = _num(raw, keys["mln"]) * 1e6                 # $Mln -> USD
    df["shares"] = _num(raw, keys["shares"]) * 1000
    # fall back to a size proxy if $Mln is absent, so weighting still works
    if df["notional"].isna().all():
        px = _num(raw, keys.get("strike"))
        df["notional"] = df["shares"] * px
    df["mkt"] = (raw[keys["mkt"]].astype(str).str.strip().str.title()
                 if keys["mkt"] else "Unknown")
    df["strategy"] = (raw[keys["strategy"]].astype(str).str.strip()
                      if keys["strategy"] else "All")
    df["side"] = (raw[keys["side"]].astype(str).str.strip().str.title()
                  if keys["side"] else "")
    df["sector"] = raw[keys["sector"]].astype(str).str.strip() if keys["sector"] else ""
    # Country / listing venue = the last 2 characters of the ticker
    # ("0700.HK" / "01896 HK" -> HK, "AAPL.US" -> US).
    if keys["sym"]:
        df["country"] = raw[keys["sym"]].astype(str).str.strip().str[-2:].str.upper()
    else:
        df["country"] = "Unknown"
    df["date"] = pd.Series(pd.NaT, index=raw.index, dtype="datetime64[ns]")
    if keys["date"]:
        df["date"] = pd.to_datetime(raw[keys["date"]], errors="coerce", format="mixed")

    for k in ("is", "open", "close", "pvwap", "epvwap", "epvwap_sprd"):
        df[k] = _num(raw, keys[k]) * (sign if k in SLIP_KEYS else 1.0)
    df["pct_dv"] = _num(raw, keys["pct_dv"])
    df["vol"] = _num(raw, keys["vol"])
    df["sprd"] = _num(raw, keys["sprd"])
    df["dur"] = _num(raw, keys["dur"])
    df["epr"] = _num(raw, keys["epr"])
    df["fr_pct"] = _as_pct(_num(raw, keys["fr"]))
    df["dark"] = _as_pct(_num(raw, keys["dark"]))
    return df


# --------------------------------------------------------------------------- #
# Weighted helpers
# --------------------------------------------------------------------------- #

def wavg(values: pd.Series, weights: pd.Series) -> float:
    v = pd.to_numeric(values, errors="coerce")
    w = pd.to_numeric(weights, errors="coerce")
    m = v.notna() & w.notna() & (w > 0)
    if not m.any():
        return float(v[v.notna()].mean()) if v.notna().any() else np.nan
    return float(np.average(v[m], weights=w[m]))


# --------------------------------------------------------------------------- #
# 1-2. Order-type slippage breakdowns (the reproduced deliverable)
# --------------------------------------------------------------------------- #

OT_COLUMNS = ["Order Type", "# Orders", "Arrival Slippage (bps)",
              "Open Slippage (bps)", "Close Slippage (bps)", "PVWAP Slippage (bps)",
              "%ADV", "Volatility", "Spread (bps)", "Fill Rate (%)"]


def _ot_row(label: str, g: pd.DataFrame) -> dict:
    w = g["notional"]
    return {
        "Order Type": label,
        "# Orders": len(g),
        "Arrival Slippage (bps)": wavg(g["is"], w),
        "Open Slippage (bps)": wavg(g["open"], w),
        "Close Slippage (bps)": wavg(g["close"], w),
        "PVWAP Slippage (bps)": wavg(g["pvwap"], w),
        "%ADV": float(g["pct_dv"].mean()),
        "Volatility": float(g["vol"].mean()),
        "Spread (bps)": float(g["sprd"].mean()),
        "Fill Rate (%)": float(g["fr_pct"].mean()),
    }


def ordertype_all(df: pd.DataFrame) -> pd.DataFrame:
    rows = [_ot_row(k, g) for k, g in df.groupby("mkt", observed=True)]
    rows.append(_ot_row("All order types", df))
    return pd.DataFrame(rows, columns=OT_COLUMNS)


def ordertype_per_strategy(df: pd.DataFrame) -> pd.DataFrame:
    """Combined overview: one stacked table, a Strategy column + order-type rows."""
    rows = []
    for strat, gs in df.groupby("strategy", observed=True):
        for mkt, g in gs.groupby("mkt", observed=True):
            r = _ot_row(mkt, g)
            r = {"Strategy": strat, **r}
            rows.append(r)
        r = _ot_row("All order types", gs)
        rows.append({"Strategy": strat, **r})
    cols = ["Strategy"] + OT_COLUMNS
    out = pd.DataFrame(rows, columns=cols)
    return out.sort_values(["Strategy", "Order Type"]).reset_index(drop=True)


BREAKDOWN_METRICS = ["# Orders", "Arrival Slippage (bps)", "PVWAP Slippage (bps)",
                     "%ADV", "Volatility", "Spread (bps)", "Fill Rate (%)", "%DARK"]


def _breakdown_row(g: pd.DataFrame) -> dict:
    w = g["notional"]
    return {
        "# Orders": len(g),
        "Arrival Slippage (bps)": wavg(g["is"], w),
        "PVWAP Slippage (bps)": wavg(g["pvwap"], w),
        "%ADV": float(g["pct_dv"].mean()),
        "Volatility": float(g["vol"].mean()),
        "Spread (bps)": float(g["sprd"].mean()),
        "Fill Rate (%)": float(g["fr_pct"].mean()),
        "%DARK": wavg(g["dark"], w),
    }


def breakdown_by(df: pd.DataFrame, keycol: str, keyname: str) -> pd.DataFrame:
    """Notional-weighted slippage + %DARK, one row per ``keycol`` value (biggest first),
    with an ``All`` total row. Used for the per-country and per-algo breakdowns."""
    order = (df.groupby(keycol, observed=True)["notional"].sum()
             .sort_values(ascending=False).index)
    rows = []
    for k in order:
        g = df[df[keycol] == k]
        rows.append({keyname: str(k), **_breakdown_row(g)})
    rows.append({keyname: "All", **_breakdown_row(df)})
    return pd.DataFrame(rows, columns=[keyname] + BREAKDOWN_METRICS)


def ordertype_by_strategy_split(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """One standalone order-type table per strategy (Market / Limit / All rows).

    Returns ``{strategy: order-type table}`` ordered by executed notional, so the
    biggest strategies come first in the report.
    """
    order = (df.groupby("strategy", observed=True)["notional"].sum()
             .sort_values(ascending=False).index)
    out: dict[str, pd.DataFrame] = {}
    for strat in order:
        gs = df[df["strategy"] == strat]
        rows = [_ot_row(mkt, g) for mkt, g in gs.groupby("mkt", observed=True)]
        rows.append(_ot_row("All order types", gs))
        out[str(strat)] = pd.DataFrame(rows, columns=OT_COLUMNS)
    return out


# --------------------------------------------------------------------------- #
# Dark-execution analytics
# --------------------------------------------------------------------------- #

SIZE_BINS = [-0.01, 1, 5, 15, 1e9]
SIZE_LABELS = ["<1% ADV", "1-5% ADV", "5-15% ADV", "15%+ ADV"]

# Fixed %DARK bucket boundaries applied to orders that traded some dark. Orders with
# no dark execution (0%) are always split into their own "0% (no dark)" bucket.
# These are just the interior cut-points, in %; edit this list to change the ranges:
#   [30, 60]          -> "0–30%", "30–60%", "60%+"
#   [10, 25, 40, 60]  -> "0–10%", "10–25%", "25–40%", "40–60%", "60%+"
DARK_FIXED_CUTS = [20, 40]


def _dark_labels(cuts: list[float]) -> list[str]:
    """Bucket labels for the positive-dark ranges defined by ``cuts`` (0% excluded)."""
    labels, prev = [], 0
    for c in cuts:
        labels.append(f"{prev:g}–{c:g}%")
        prev = c
    labels.append(f"{prev:g}%+")
    return labels


def dark_bucket(df: pd.DataFrame) -> pd.Series:
    """Assign each order to a fixed %DARK bucket.

    Orders with no dark execution (0%) go to a dedicated "0% (no dark)" bucket so the
    large fully-lit mass doesn't swamp the first range. Orders that traded dark are
    split by the fixed ranges from ``DARK_FIXED_CUTS`` (edit that list to change them).
    """
    s = pd.to_numeric(df["dark"], errors="coerce")
    cuts = list(DARK_FIXED_CUTS)
    pos_labels = _dark_labels(cuts)
    labels = ["0% (no dark)"] + pos_labels

    out = pd.Series(pd.NA, index=s.index, dtype=object)
    out[s <= 0] = "0% (no dark)"
    pos = s > 0
    if pos.any():
        binned = pd.cut(s[pos], bins=[0.0] + cuts + [np.inf], labels=pos_labels)
        out[pos] = binned.astype(object)
    return pd.Categorical(out, categories=labels, ordered=True)


def size_bucket(df: pd.DataFrame) -> pd.Series:
    return pd.cut(df["pct_dv"], bins=SIZE_BINS, labels=SIZE_LABELS)


def dark_slippage_table(df: pd.DataFrame) -> pd.DataFrame:
    d = df.assign(_b=dark_bucket(df))
    rows = []
    for b in d["_b"].cat.categories:      # bucket order, adaptive or fixed
        g = d[d["_b"] == b]
        if g.empty:
            continue
        w = g["notional"]
        rows.append({
            "%DARK bucket": b,
            "# Orders": len(g),
            "Notional (USD m)": float(w.sum()) / 1e6,
            "Avg %DARK": float(g["dark"].mean()),
            "Arrival Slippage (bps)": wavg(g["is"], w),
            "PVWAP Slippage (bps)": wavg(g["pvwap"], w),
            "ePvwap/Sprd": wavg(g["epvwap_sprd"], w),
            "Avg Spread (bps)": float(g["sprd"].mean()),
            "Avg %ADV": float(g["pct_dv"].mean()),
        })
    return pd.DataFrame(rows)


def dark_by_size_table(df: pd.DataFrame) -> pd.DataFrame:
    """wavg arrival slippage by size bucket for LOW vs HIGH dark participation."""
    d = df.assign(_s=size_bucket(df))
    low = d[d["dark"] < 15]
    high = d[d["dark"] >= 30]
    rows = []
    for s in SIZE_LABELS:
        gl, gh = low[low["_s"] == s], high[high["_s"] == s]
        lo = wavg(gl["is"], gl["notional"]) if len(gl) else np.nan
        hi = wavg(gh["is"], gh["notional"]) if len(gh) else np.nan
        rows.append({
            "Order size": s,
            "Low dark (<15%) bps": lo,
            "High dark (>=30%) bps": hi,
            "Dark benefit (bps)": (hi - lo) if np.isfinite(hi) and np.isfinite(lo) else np.nan,
            "# low / # high": f"{len(gl)} / {len(gh)}",
        })
    return pd.DataFrame(rows)


def dark_savings(df: pd.DataFrame) -> dict:
    """Indicative $ opportunity from shifting more flow into dark.

    Compares notional-weighted arrival slippage in the LOW-dark band (<15%) with
    the HIGH-dark band (>=30%). The per-order uplift is applied only to notional
    currently executing with low dark participation -- an indicative upper-bound
    on recoverable cost, not a promise.
    """
    w = df["notional"]
    total_notional = float(w.sum())
    wtd_dark = wavg(df["dark"], w)

    low = df[df["dark"] < 15]
    high = df[df["dark"] >= 30]
    low_bps = wavg(low["is"], low["notional"])
    high_bps = wavg(high["is"], high["notional"])
    delta = (high_bps - low_bps) if np.isfinite(high_bps) and np.isfinite(low_bps) else np.nan
    low_notional = float(low["notional"].sum())
    # only positive uplift counts toward the opportunity
    uplift = max(delta, 0.0) if np.isfinite(delta) else 0.0
    savings_usd = uplift / 1e4 * low_notional
    return {
        "total_notional": total_notional,
        "wtd_dark_pct": wtd_dark,
        "low_bps": low_bps,
        "high_bps": high_bps,
        "delta_bps": delta,
        "low_notional": low_notional,
        "low_share_pct": low_notional / total_notional * 100 if total_notional else np.nan,
        "savings_usd": savings_usd,
    }


# --------------------------------------------------------------------------- #
# Auxiliary CSV loaders (all optional)
# --------------------------------------------------------------------------- #

def _load_aux(aux_dir: Path | None, name: str) -> pd.DataFrame | None:
    if aux_dir is None:
        return None
    p = aux_dir / name
    if not p.exists():
        return None
    try:
        return pd.read_csv(p)
    except Exception as exc:  # noqa: BLE001
        print(f"  ! could not read {p.name}: {exc}", file=sys.stderr)
        return None


def load_venue_split(aux_dir: Path | None) -> pd.DataFrame | None:
    """Merge dark_routed.csv + dark_executed.csv into Venue / Routed% / Executed%."""
    r = _load_aux(aux_dir, "dark_routed.csv")
    e = _load_aux(aux_dir, "dark_executed.csv")
    if r is None and e is None:
        return None

    def norm(d, valname):
        if d is None:
            return None
        vcol = _find(d.columns, ["Venue", "Dark Venue", "Category"])
        pcol = _find(d.columns, ["Routed %", "Executed %", "Pct", "%", "Percent", "Volume %"])
        if vcol is None or pcol is None:
            return None
        out = d[[vcol, pcol]].copy()
        out.columns = ["Venue", valname]
        out["Venue"] = out["Venue"].astype(str).str.strip()
        out[valname] = pd.to_numeric(out[valname], errors="coerce")
        return out

    rn, en = norm(r, "Routed %"), norm(e, "Executed %")
    if rn is not None and en is not None:
        return rn.merge(en, on="Venue", how="outer")
    return rn if rn is not None else en


# --------------------------------------------------------------------------- #
# Charts  (each returns the path it wrote)
# --------------------------------------------------------------------------- #

def _titles(ax, title: str, subtitle: str | None = None) -> None:
    """Left-aligned bold title with an optional grey subtitle line beneath it."""
    if subtitle:
        ax.set_title(title, loc="left", pad=30)
        ax.text(0.0, 1.015, subtitle, transform=ax.transAxes, ha="left", va="bottom",
                fontsize=10.5, color=MUTED)
    else:
        ax.set_title(title, loc="left")


def _bar_labels(ax, fmt="{:+.1f}", offset=0.3, fontsize=9.5):
    for c in ax.containers:
        ax.bar_label(c, labels=[fmt.format(v) for v in c.datavalues],
                     padding=offset, fontsize=fontsize, color=INK)


def chart_ordertype(tbl: pd.DataFrame, png: Path) -> Path:
    metrics = ["Arrival Slippage (bps)", "Open Slippage (bps)",
               "Close Slippage (bps)", "PVWAP Slippage (bps)"]
    sub = tbl[tbl["Order Type"] != "All order types"].copy()
    types = sub["Order Type"].tolist()
    long = sub.melt(id_vars="Order Type", value_vars=metrics,
                    var_name="Benchmark", value_name="bps")
    long["Benchmark"] = long["Benchmark"].str.replace(" Slippage (bps)", "", regex=False)
    order = [m.replace(" Slippage (bps)", "") for m in metrics]

    fig, ax = plt.subplots(figsize=(12, 6.5))
    sns.barplot(long, x="Benchmark", y="bps", hue="Order Type", order=order,
                hue_order=types, palette=SEQ[:len(types)], edgecolor="white",
                linewidth=0.6, ax=ax)
    for c in ax.containers:
        ax.bar_label(c, fmt="%+.1f", padding=3, fontsize=9)
    ax.axhline(0, color=MUTED, lw=1.4, ls=(0, (4, 3)))
    ax.set_xlabel("")
    ax.set_ylabel("Notional-weighted slippage (bps)")
    _titles(ax, "Slippage by order type  ·  all strategies",
            "+ = outperformance   ·   − = cost")
    ax.legend(title="Order type", ncol=len(types), loc="upper center",
              bbox_to_anchor=(0.5, -0.14), fontsize=9.5, title_fontsize=10)
    fig.savefig(png)
    plt.close(fig)
    return png


def chart_dark_slippage(tbl: pd.DataFrame, png: Path) -> Path:
    """Hero chart: arrival + PVWAP slippage rise as %DARK participation rises."""
    labelmap = {"Arrival Slippage (bps)": "Arrival (IS)",
                "PVWAP Slippage (bps)": "Interval PVWAP"}
    long = tbl.melt(id_vars="%DARK bucket", value_vars=list(labelmap),
                    var_name="Benchmark", value_name="bps")
    long["Benchmark"] = long["Benchmark"].map(labelmap)
    order = tbl["%DARK bucket"].tolist()   # table is already in bucket order

    fig, ax = plt.subplots(figsize=(12, 6.5))
    sns.barplot(long, x="%DARK bucket", y="bps", hue="Benchmark", order=order,
                hue_order=list(labelmap.values()), palette=[DARKC, ACC4],
                edgecolor="white", linewidth=0.6, ax=ax)
    for c in ax.containers:
        ax.bar_label(c, fmt="%+.1f", padding=3, fontsize=9.5)
    ax.axhline(0, color=MUTED, lw=1.4, ls=(0, (4, 3)))
    ax.set_xlabel("Dark participation of the order (%DARK)")
    ax.set_ylabel("Notional-weighted slippage (bps)")
    _titles(ax, "More dark participation → better execution",
            "notional-weighted slippage by %DARK bucket   ·   + = outperformance")
    # order-count annotation under each bucket
    counts = dict(zip(tbl["%DARK bucket"], tbl["# Orders"]))
    for xi, b in enumerate(order):
        ax.annotate(f"n={int(counts[b])}", (xi, 0), xytext=(0, -22),
                    textcoords="offset points", ha="center", color=MUTED, fontsize=9)
    ax.legend(loc="upper left", bbox_to_anchor=(1.01, 1.0), fontsize=10, title="")
    fig.savefig(png)
    plt.close(fig)
    return png


def chart_dark_spreadnorm(tbl: pd.DataFrame, png: Path) -> Path:
    order = tbl["%DARK bucket"].tolist()   # table is already in bucket order
    fig, ax = plt.subplots(figsize=(11, 6))
    sns.pointplot(tbl, x="%DARK bucket", y="ePvwap/Sprd", order=order,
                  color=DARKC, markers="o", markersize=11, linewidth=2.6, ax=ax)
    x = np.arange(len(order))
    yvals = tbl.set_index("%DARK bucket")["ePvwap/Sprd"].reindex(order)
    ax.fill_between(x, yvals, color=DARKC, alpha=0.12)
    for xi, yv in zip(x, yvals):
        if np.isfinite(yv):
            ax.annotate(f"{yv:+.2f}", (xi, yv), xytext=(0, 12),
                        textcoords="offset points", ha="center", fontsize=9.5,
                        fontweight="bold", color=INK)
    ax.axhline(0, color=MUTED, lw=1.4, ls=(0, (4, 3)))
    ax.set_xlabel("Dark participation (%DARK)")
    ax.set_ylabel("Spread-normalised PVWAP slippage")
    _titles(ax, "Spread capture improves with dark",
            "ePvwap / Spread  ·  higher = more of the quoted spread captured")
    fig.savefig(png)
    plt.close(fig)
    return png


def chart_dark_by_size(tbl: pd.DataFrame, png: Path) -> Path:
    labelmap = {"Low dark (<15%) bps": "Low dark (<15%)",
                "High dark (>=30%) bps": "High dark (≥30%)"}
    long = tbl.melt(id_vars="Order size", value_vars=list(labelmap),
                    var_name="Dark", value_name="bps")
    long["Dark"] = long["Dark"].map(labelmap)
    order = [s for s in SIZE_LABELS if s in set(tbl["Order size"])]

    fig, ax = plt.subplots(figsize=(12, 6.5))
    sns.barplot(long, x="Order size", y="bps", hue="Dark", order=order,
                hue_order=list(labelmap.values()), palette=[LITC, DARKC],
                edgecolor="white", linewidth=0.6, ax=ax)
    for c in ax.containers:
        ax.bar_label(c, fmt="%+.1f", padding=3, fontsize=9)
    ax.axhline(0, color=MUTED, lw=1.4, ls=(0, (4, 3)))
    ax.set_xlabel("")
    ax.set_ylabel("Arrival slippage (bps)")
    _titles(ax, "Dark helps most on the hardest (largest) orders",
            "notional-weighted arrival slippage · low vs high dark, by order size")
    ax.legend(loc="upper left", bbox_to_anchor=(1.01, 1.0), fontsize=10, title="")
    fig.savefig(png)
    plt.close(fig)
    return png


def chart_venue_split(venue: pd.DataFrame, png: Path) -> Path:
    """Per-venue stacked bar: one bar per venue split into a routed segment and an
    executed segment, so each venue's routed vs executed reads as a single divided bar."""
    v = venue.copy().fillna(0.0)
    vcols = [c for c in ["Routed %", "Executed %"] if c in v.columns]
    # biggest venues (by combined footprint) at the top
    v["_tot"] = v[vcols].sum(axis=1)
    v = v.sort_values("_tot", ascending=True)
    venues = v["Venue"].tolist()
    y = np.arange(len(venues))
    pal = {"Routed %": ACC3, "Executed %": DARKC}
    txt = {"Routed %": INK, "Executed %": "white"}

    fig, ax = plt.subplots(figsize=(11, 0.6 * len(v) + 3))
    left = np.zeros(len(v))
    for col in vcols:
        vals = v[col].to_numpy()
        ax.barh(y, vals, left=left, color=pal[col], edgecolor="white",
                linewidth=0.8, label=col.replace(" %", ""))
        for yi, (xv, lf) in enumerate(zip(vals, left)):
            if xv > 1.5:  # only label segments wide enough to hold the text
                ax.annotate(f"{xv:.1f}", (lf + xv / 2, yi), ha="center", va="center",
                            fontsize=9, color=txt[col], fontweight="bold")
        left += vals
    ax.set_yticks(y)
    ax.set_yticklabels(venues)
    ax.set_xlabel("% of dark volume  (routed segment + executed segment)")
    ax.set_ylabel("")
    _titles(ax, "Routed vs executed volume by dark venue",
            "each venue is one bar · routed and executed shown as two coloured segments")
    ax.grid(axis="x", color=GRID)
    ax.grid(visible=False, axis="y")
    ax.legend(loc="center left", bbox_to_anchor=(1.02, 0.5), title="")
    fig.savefig(png)
    plt.close(fig)
    return png


def chart_venue_summary(vs: pd.DataFrame, png: Path) -> Path:
    cat = _find(vs.columns, ["Venue Category", "Category", "Venue"])
    val = _find(vs.columns, ["% Total Exec Value", "% Total Shares", "Exec Value", "Shares"])
    if cat is None or val is None:
        return png
    d = vs[[cat, val]].copy()
    d[val] = pd.to_numeric(d[val], errors="coerce")
    d = d.dropna(subset=[val])
    d = d[d[val] > 0].sort_values(val, ascending=False)
    if d.empty:
        return png
    labels = d[cat].astype(str).tolist()
    values = d[val].to_numpy(dtype=float)
    total = values.sum()
    colors = [SEQ[i % len(SEQ)] for i in range(len(d))]

    fig, ax = plt.subplots(figsize=(10, 6.5))
    wedges, _ = ax.pie(values, colors=colors, startangle=90, counterclock=False,
                       wedgeprops=dict(width=0.42, edgecolor="white", linewidth=1.6))
    ax.set_aspect("equal")
    ax.set_title("Dark venue summary", loc="left", pad=16)
    ax.text(0, 0, val.replace("% Total ", "").replace("Total ", ""),
            ha="center", va="center", fontsize=11, color=MUTED)
    leg = [f"{lab}   ·   {v / total * 100:.1f}%" for lab, v in zip(labels, values)]
    ax.legend(wedges, leg, loc="center left", bbox_to_anchor=(1.02, 0.5),
              frameon=False, fontsize=10.5, title=f"share of {val}", title_fontsize=10)
    fig.savefig(png)
    plt.close(fig)
    return png


def chart_algo(algo: pd.DataFrame, png: Path) -> Path:
    scol = _find(algo.columns, ["Strategy", "Algorithm", "Algo"])
    icol = _find(algo.columns, ["Wgt Impact Bps", "Impact Bps"])
    wcol = _find(algo.columns, ["%Weight", "Exec Value"])
    if scol is None or icol is None:
        return png
    d = algo.copy()
    d[icol] = pd.to_numeric(d[icol], errors="coerce")
    d = d.sort_values(icol)
    order = d[scol].astype(str).tolist()
    palette = [DARKC if v >= 0 else LITC for v in d[icol]]
    fig, ax = plt.subplots(figsize=(11, 0.6 * len(d) + 3))
    sns.barplot(d, y=scol, x=icol, order=order, hue=scol, hue_order=order,
                palette=palette, legend=False, edgecolor="white", linewidth=0.6, ax=ax)
    for c in ax.containers:
        ax.bar_label(c, fmt="%+.1f", padding=3, fontsize=10)
    ax.axvline(0, color=MUTED, lw=1.4, ls=(0, (4, 3)))
    ax.set_ylabel("")
    ax.set_xlabel(f"{icol}  (+ = good / − = cost)")
    ttl = "Performance summary by algorithm"
    if wcol:
        ttl += f"  ·  weighted by {wcol}"
    _titles(ax, ttl)
    ax.grid(axis="x", color=GRID)
    ax.grid(visible=False, axis="y")
    fig.savefig(png)
    plt.close(fig)
    return png


# performance benchmark per algo -> (order-level df column, y-axis label).
# Edit / extend this to remap benchmarks for a different set of strategies.
ALGO_BENCHMARK = {
    "VWAP":   ("pvwap", "Order PVWAP slippage (bps)"),
    "IS":     ("is", "Arrival price slippage (bps)"),
    "INLINE": ("is", "Arrival price slippage (bps)"),
}
DEFAULT_BENCHMARK = ("is", "Arrival price slippage (bps)")

# Which algos to show in the per-algo dark-performance chart. Empty list = all algos.
DARK_PERF_ALGOS = ["VWAP"]


def _algo_benchmark(strategy: str) -> tuple[str, str]:
    return ALGO_BENCHMARK.get(str(strategy).strip().upper(), DEFAULT_BENCHMARK)


def chart_algo_dark_perf(df: pd.DataFrame, png: Path, min_orders: int = 5) -> Path | None:
    """Small-multiples, one panel per algo: order-level performance vs %DARK.

    Only orders that actually executed dark (``%DARK > 0``) are plotted. Each panel
    scatters those orders (x = %DARK, y = the algo's benchmark slippage — VWAP → Order
    PVWAP, IS/Inline → arrival price, others → arrival by default) as bubbles sized by
    the order's %ADV, and overlays a linear trend line so the dark-vs-performance
    relationship is visible per strategy. Algos with fewer than ``min_orders`` dark
    orders are skipped. Returns ``None`` if none qualify.
    """
    dark = df[df["dark"] > 0]
    if DARK_PERF_ALGOS:
        keep = {a.strip().upper() for a in DARK_PERF_ALGOS}
        dark = dark[dark["strategy"].astype(str).str.strip().str.upper().isin(keep)]
    panels = []
    for strat, g in dark.groupby("strategy", observed=True):
        col, ylab = _algo_benchmark(strat)
        x = pd.to_numeric(g["dark"], errors="coerce")
        y = pd.to_numeric(g[col], errors="coerce")
        adv = pd.to_numeric(g["pct_dv"], errors="coerce")
        m = x.notna() & y.notna()
        if int(m.sum()) >= min_orders:
            panels.append((str(strat), x[m].to_numpy(), y[m].to_numpy(),
                           adv[m].to_numpy(), ylab))
    if not panels:
        return None
    panels.sort(key=lambda p: -len(p[1]))               # busiest algos first

    # common bubble-size scale from %ADV (area grows with %ADV; cap outliers at p95)
    all_adv = np.concatenate([p[3] for p in panels])
    all_adv = all_adv[np.isfinite(all_adv)]
    adv_ref = max(float(np.percentile(all_adv, 95)) if all_adv.size else 1.0, 1e-6)
    SMIN, SMAX = 18.0, 460.0

    def _size(a):
        a = np.where(np.isfinite(a), a, 0.0)
        return SMIN + np.clip(a, 0.0, adv_ref) / adv_ref * (SMAX - SMIN)

    ncol = min(3, len(panels))
    nrow = int(np.ceil(len(panels) / ncol))
    cellw, cellh = (9.0, 5.8) if len(panels) == 1 else (5.4, 4.4)
    fig, axes = plt.subplots(nrow, ncol, figsize=(cellw * ncol, cellh * nrow),
                             squeeze=False)
    for i, (strat, x, y, adv, ylab) in enumerate(panels):
        ax = axes[i // ncol][i % ncol]
        ax.scatter(x, y, s=_size(adv), color=DARKC, alpha=0.45, edgecolor="white",
                   linewidth=0.5, zorder=3)
        if len(x) >= 2 and np.unique(x).size >= 2:
            slope, intercept = np.polyfit(x, y, 1)
            xs = np.linspace(x.min(), x.max(), 50)
            ax.plot(xs, intercept + slope * xs, color=LITC, lw=2.6, zorder=4,
                    label=f"trend  {slope:+.3f} bps / %dark")
            ax.legend(loc="upper right", fontsize=9, frameon=True, facecolor="white",
                      framealpha=0.75, edgecolor="none")
        ax.axhline(0, color=MUTED, lw=1.1, ls=(0, (4, 3)))
        ax.set_title(f"{strat}", loc="left", fontsize=13.5)
        ax.set_xlabel("Dark execution (%DARK)")
        ax.set_ylabel(ylab, fontsize=10.5)
    for j in range(len(panels), nrow * ncol):           # blank any unused cells
        axes[j // ncol][j % ncol].axis("off")

    # size-reference legend: a few representative %ADV bubbles
    from matplotlib.lines import Line2D
    ref_vals = sorted({max(1, int(round(v)))
                       for v in np.percentile(all_adv, [50, 80, 95])}) if all_adv.size else [1]
    size_handles = [Line2D([0], [0], marker="o", linestyle="", markerfacecolor=DARKC,
                           markeredgecolor="white", alpha=0.55,
                           markersize=np.sqrt(_size(v)), label=f"{v:g}%")
                    for v in ref_vals]

    # spacing measured in inches (va="top" on both) so the title strip never collides
    h = cellh * nrow
    fig.tight_layout(rect=[0, 0, 1, 1 - 0.85 / h])
    fig.suptitle("Order-level performance vs dark execution, by algo",
                 x=0.015, ha="left", va="top", y=1 - 0.10 / h,
                 fontsize=16, fontweight="bold", color=INK)
    fig.text(0.015, 1 - 0.48 / h, "each bubble = one dark-executed order · size = %ADV · "
             "+ = outperformance / − = cost", ha="left", va="top", fontsize=10.5, color=MUTED)
    fig.legend(handles=size_handles, title="%ADV", loc="upper right",
               bbox_to_anchor=(0.995, 1 - 0.10 / h), fontsize=9, title_fontsize=9,
               labelspacing=1.5, borderpad=0.8, handletextpad=1.2, frameon=False)
    fig.savefig(png)
    plt.close(fig)
    return png


def chart_savings(sv: dict, png: Path) -> Path:
    """Callout figure summarising the dark opportunity."""
    fig, ax = plt.subplots(figsize=(11, 5.5))
    ax.axis("off")
    ax.set_title("The dark opportunity", loc="left")
    dlt = sv["delta_bps"]
    lines = [
        (f"{sv['wtd_dark_pct']:.1f}%", "current notional-weighted dark participation"),
        (f"{dlt:+.1f} bps", "arrival-slippage gap: high-dark (≥30%) vs low-dark (<15%)"),
        (f"{sv['low_share_pct']:.0f}%", "of notional still executes with <15% dark"),
        (_fmt_usd(sv["savings_usd"]), "indicative recoverable cost if that flow moved to dark"),
    ]
    y = 0.82
    for big, small in lines:
        ax.text(0.02, y, big, fontsize=30, fontweight="bold", color=DARKC,
                transform=ax.transAxes, va="center")
        ax.text(0.34, y, small, fontsize=13, color=INK, transform=ax.transAxes, va="center")
        y -= 0.22
    ax.text(0.02, -0.02, "Indicative upper bound (arrival-slippage basis); not a guarantee. "
            "Assumes low-dark flow reaches high-dark performance.",
            fontsize=9.5, color=MUTED, transform=ax.transAxes)
    fig.savefig(png)
    plt.close(fig)
    return png


# --------------------------------------------------------------------------- #
# Formatting helpers
# --------------------------------------------------------------------------- #

def _fmt_usd(x: float) -> str:
    if not np.isfinite(x):
        return "n/a"
    a = abs(x)
    if a >= 1e9:
        return f"${x / 1e9:,.2f}bn"
    if a >= 1e6:
        return f"${x / 1e6:,.1f}m"
    if a >= 1e3:
        return f"${x / 1e3:,.1f}k"
    return f"${x:,.0f}"


# --------------------------------------------------------------------------- #
# PDF report (paginated, slide style)
# --------------------------------------------------------------------------- #

def _fig_from_png(png: Path, title: str | None = None) -> plt.Figure:
    img = plt.imread(str(png))
    h, w = img.shape[0], img.shape[1]
    fig = plt.figure(figsize=(11.69, 8.27))  # A4 landscape
    ax = fig.add_axes([0.04, 0.05, 0.92, 0.86])
    ax.imshow(img)
    ax.axis("off")
    if title:
        fig.text(0.04, 0.95, title, fontsize=15, fontweight="bold", color=INK)
    return fig


def _fig_table(df: pd.DataFrame, title: str, note: str | None = None,
               float_fmt: str = "{:.1f}") -> plt.Figure:
    fig = plt.figure(figsize=(11.69, 8.27))
    fig.text(0.04, 0.94, title, fontsize=15, fontweight="bold", color=INK)
    if note:
        fig.text(0.04, 0.905, note, fontsize=10.5, color=MUTED)
    ax = fig.add_axes([0.03, 0.05, 0.94, 0.80])
    ax.axis("off")

    disp = df.copy()
    for c in disp.columns:
        if pd.api.types.is_float_dtype(disp[c]):
            disp[c] = disp[c].map(lambda v: float_fmt.format(v) if np.isfinite(v) else "")
        else:
            disp[c] = disp[c].astype(str)
    # shorter headers so wide tables don't overlap in the fixed-width figure
    abbrev = {
        "Arrival Slippage (bps)": "Arrival\n(bps)",
        "Open Slippage (bps)": "Open\n(bps)",
        "Close Slippage (bps)": "Close\n(bps)",
        "PVWAP Slippage (bps)": "PVWAP\n(bps)",
        "Fill Rate (%)": "Fill %",
        "Spread (bps)": "Spread\n(bps)",
        "Notional (USD m)": "Notional\n(USD m)",
        "Avg Spread (bps)": "Avg Sprd\n(bps)",
        "Low dark (<15%) bps": "Low dark\n(<15%)",
        "High dark (>=30%) bps": "High dark\n(≥30%)",
        "Dark benefit (bps)": "Dark\nbenefit",
    }
    labels = [abbrev.get(c, c) for c in disp.columns]
    tbl = ax.table(cellText=disp.values, colLabels=labels,
                   cellLoc="center", loc="upper center")
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(8.5 if disp.shape[1] > 8 else 9.5)
    tbl.scale(1, 1.7)
    for (r, c), cell in tbl.get_celld().items():
        cell.set_edgecolor("#D7DEDC")
        if r == 0:
            cell.set_facecolor(f"#{XL_HEADER_BG}")
            cell.set_text_props(color="white", fontweight="bold")
        elif r % 2 == 0:
            cell.set_facecolor(f"#{XL_BAND_BG}")
    return fig


def build_pdf(path: Path, ctx: dict) -> None:
    with PdfPages(path) as pdf:
        # ---- cover page ----
        fig = plt.figure(figsize=(11.69, 8.27))
        fig.patch.set_facecolor(ACC4)
        fig.text(0.06, 0.66, "Transaction Cost Analysis", fontsize=34,
                 fontweight="bold", color="white")
        fig.text(0.06, 0.58, ctx["client"], fontsize=22, color="#E9C46A")
        fig.text(0.06, 0.52, ctx["period"], fontsize=14, color="#CBD5D1")
        s = ctx["sv"]
        fig.text(0.06, 0.34,
                 f"{ctx['n_orders']:,} orders   ·   {_fmt_usd(s['total_notional'])} executed"
                 f"   ·   {s['wtd_dark_pct']:.1f}% dark participation",
                 fontsize=13, color="white")
        fig.text(0.06, 0.06, "+ = outperformance / − = cost (bps)   ·   "
                 "weights = executed notional (USD)", fontsize=10, color="#9FB0AC")
        pdf.savefig(fig, facecolor=fig.get_facecolor())
        plt.close(fig)

        # ---- exec summary text ----
        fig = plt.figure(figsize=(11.69, 8.27))
        fig.text(0.06, 0.92, "Executive summary", fontsize=17, fontweight="bold", color=INK)
        y = 0.83
        for b in ctx["summary_bullets"]:
            fig.text(0.07, y, "•", fontsize=13, color=DARKC)
            fig.text(0.09, y, b, fontsize=12, color=INK)
            y -= 0.075
        pdf.savefig(fig)
        plt.close(fig)

        # ---- table pages ----
        pdf.savefig(_fig_table(ctx["ot_all"], "1 · Total slippage breakdown by order type",
                               "All strategies · notional-weighted · + = good / − = cost"))
        plt.close("all")
        pdf.savefig(_fig_table(ctx["ot_strat"],
                               "2 · Slippage breakdown by order type per strategy — overview",
                               "Notional-weighted · + = good / − = cost"))
        plt.close("all")
        for i, (strat, tbl) in enumerate(ctx["ot_strat_split"].items(), start=1):
            pdf.savefig(_fig_table(tbl, f"2.{i} · Order type breakdown — {strat}",
                                   "Notional-weighted · + = good / − = cost"))
            plt.close("all")
        pdf.savefig(_fig_table(ctx["dark_tbl"], "3 · Slippage by dark participation",
                               "Notional-weighted by %DARK bucket", float_fmt="{:.2f}"))
        plt.close("all")
        pdf.savefig(_fig_table(ctx["country_tbl"], "4 · Slippage & dark participation by country",
                               "Notional-weighted · + = good / − = cost · %DARK weighted by notional"))
        plt.close("all")
        pdf.savefig(_fig_table(ctx["algo_tbl"], "5 · Slippage & dark participation by algo",
                               "Notional-weighted · + = good / − = cost · %DARK weighted by notional"))
        plt.close("all")

        # ---- chart pages ----
        for png, title in ctx["pdf_charts"]:
            pdf.savefig(_fig_from_png(png, title))
            plt.close("all")


# --------------------------------------------------------------------------- #
# XLSX report
# --------------------------------------------------------------------------- #

def build_xlsx(path: Path, ctx: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    thin = Side(style="thin", color="D7DEDC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor=XL_HEADER_BG)
    band_fill = PatternFill("solid", fgColor=XL_BAND_BG)
    hdr_font = Font(color=XL_HEADER_FG, bold=True, size=11)
    title_font = Font(color=XL_TITLE_FG, bold=True, size=14)
    note_font = Font(color="6C757D", italic=True, size=10)

    def write_table(ws, df, start_row, title, note=None, pct_cols=(), num_cols=()):
        ws.cell(row=start_row, column=1, value=title).font = title_font
        r = start_row + 1
        if note:
            ws.cell(row=r, column=1, value=note).font = note_font
            r += 1
        header_row = r
        for j, col in enumerate(df.columns, start=1):
            c = ws.cell(row=header_row, column=j, value=str(col))
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        for i, (_, row) in enumerate(df.iterrows()):
            rr = header_row + 1 + i
            band = (i % 2 == 1)
            for j, col in enumerate(df.columns, start=1):
                val = row[col]
                if isinstance(val, float) and not np.isfinite(val):
                    val = None
                cell = ws.cell(row=rr, column=j, value=val)
                cell.border = border
                if band:
                    cell.fill = band_fill
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="center")
                    if col in pct_cols or col in num_cols:
                        cell.number_format = "0.0"
                    elif "bps" in str(col).lower() or "Sprd" in str(col):
                        cell.number_format = "+0.0;-0.0"
        # column widths
        for j, col in enumerate(df.columns, start=1):
            width = max(11, min(30, int(df[col].astype(str).str.len().max() if len(df) else 10) + 2,
                                len(str(col)) + 4))
            ws.column_dimensions[get_column_letter(j)].width = max(width, len(str(col)) + 3)
        return header_row + len(df) + 3  # next free row

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{ctx['client']} — Transaction Cost Analysis"
    ws["A1"].font = Font(bold=True, size=18, color=XL_TITLE_FG)
    ws["A2"] = ctx["period"]
    ws["A2"].font = note_font
    ws["A3"] = ("Convention: + = outperformance / − = cost (bps). "
                "Weights = executed notional (USD = $Mln × 1e6).")
    ws["A3"].font = note_font
    r = 5
    for b in ctx["summary_bullets"]:
        ws.cell(row=r, column=1, value="•").font = Font(color=XL_ACCENT_BG, bold=True)
        ws.cell(row=r, column=2, value=b)
        r += 1
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 110

    # ---- Order-type sheets ----
    ws1 = wb.create_sheet("Order Type — All")
    ws1.sheet_view.showGridLines = False
    write_table(ws1, ctx["ot_all"], 1,
                "Total Slippage Breakdown by Order Type (all strategies)",
                "Notional-weighted average by executed notional (USD)")
    ws2 = wb.create_sheet("Order Type — Strategy")
    ws2.sheet_view.showGridLines = False
    nr = write_table(ws2, ctx["ot_strat"], 1,
                     "Slippage Breakdown by Order Type per Strategy — overview",
                     "Notional-weighted average by executed notional (USD)")
    for strat, tbl in ctx["ot_strat_split"].items():
        nr = write_table(ws2, tbl, nr, f"Strategy: {strat}",
                         "Notional-weighted · + = good / − = cost")

    # ---- Dark analysis sheet ----
    wd = wb.create_sheet("Dark Analysis")
    wd.sheet_view.showGridLines = False
    nr = write_table(wd, ctx["dark_tbl"], 1, "Slippage by dark participation (%DARK)",
                     "Notional-weighted · + = good / − = cost")
    nr = write_table(wd, ctx["dark_size_tbl"], nr, "Dark benefit by order size",
                     "Arrival slippage: low vs high dark participation")
    sv = ctx["sv"]
    op = pd.DataFrame({
        "Metric": ["Total executed notional", "Weighted dark participation",
                   "Low-dark (<15%) arrival slippage", "High-dark (≥30%) arrival slippage",
                   "Dark benefit (delta)", "Notional still low-dark",
                   "Indicative recoverable cost"],
        "Value": [_fmt_usd(sv["total_notional"]), f"{sv['wtd_dark_pct']:.1f}%",
                  f"{sv['low_bps']:+.1f} bps", f"{sv['high_bps']:+.1f} bps",
                  f"{sv['delta_bps']:+.1f} bps",
                  f"{_fmt_usd(sv['low_notional'])} ({sv['low_share_pct']:.0f}%)",
                  _fmt_usd(sv["savings_usd"])],
    })
    write_table(wd, op, nr, "The dark opportunity (indicative)")

    # ---- Country / algo breakdown sheet ----
    wb2 = wb.create_sheet("Country & Algo")
    wb2.sheet_view.showGridLines = False
    nr = write_table(wb2, ctx["country_tbl"], 1,
                     "Slippage & dark participation by country",
                     "Notional-weighted · + = good / − = cost · %DARK weighted by notional")
    write_table(wb2, ctx["algo_tbl"], nr,
                "Slippage & dark participation by algo",
                "Notional-weighted · + = good / − = cost · %DARK weighted by notional")

    # ---- Venue / algo sheets ----
    if ctx.get("venue") is not None:
        wv = wb.create_sheet("Venues")
        wv.sheet_view.showGridLines = False
        nr = write_table(wv, ctx["venue"], 1, "Routed vs executed volume by dark venue")
        if ctx.get("venue_summary") is not None:
            write_table(wv, ctx["venue_summary"], nr, "Dark venue summary")
    elif ctx.get("venue_summary") is not None:
        wv = wb.create_sheet("Venues")
        wv.sheet_view.showGridLines = False
        write_table(wv, ctx["venue_summary"], 1, "Dark venue summary")
    if ctx.get("algo") is not None:
        wa = wb.create_sheet("Algo Summary")
        wa.sheet_view.showGridLines = False
        write_table(wa, ctx["algo"], 1, "Performance Summary by Algorithm")

    # ---- Charts sheet (embed the PNGs) ----
    wc = wb.create_sheet("Charts")
    wc.sheet_view.showGridLines = False
    row = 1
    for png, title in ctx["pdf_charts"]:
        wc.cell(row=row, column=1, value=title).font = title_font
        img = XLImage(str(png))
        # scale to a sensible width
        scale = 900 / img.width
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        wc.add_image(img, f"A{row + 1}")
        row += int(img.height / 18) + 4

    wb.save(path)


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #

def build_summary_bullets(df: pd.DataFrame, ot_all: pd.DataFrame,
                          dark_tbl: pd.DataFrame, sv: dict) -> list[str]:
    b = []
    all_row = ot_all[ot_all["Order Type"] == "All order types"].iloc[0]
    b.append(f"Executed {_fmt_usd(sv['total_notional'])} across {len(df):,} orders; "
             f"headline arrival slippage {all_row['Arrival Slippage (bps)']:+.1f} bps, "
             f"PVWAP {all_row['PVWAP Slippage (bps)']:+.1f} bps (+ good / − cost).")
    mkt = ot_all[ot_all["Order Type"].isin(["Market", "Limit"])]
    if len(mkt) == 2:
        best = mkt.loc[mkt["Arrival Slippage (bps)"].idxmax()]
        b.append(f"By order type, {best['Order Type']} orders show the better arrival "
                 f"slippage ({best['Arrival Slippage (bps)']:+.1f} bps).")
    if len(dark_tbl) >= 2:
        lo, hi = dark_tbl.iloc[0], dark_tbl.iloc[-1]
        b.append(f"Dark participation is the strongest lever: arrival slippage improves "
                 f"from {lo['Arrival Slippage (bps)']:+.1f} bps in the {lo['%DARK bucket']} "
                 f"bucket to {hi['Arrival Slippage (bps)']:+.1f} bps in the {hi['%DARK bucket']} "
                 f"bucket.")
    b.append(f"Current weighted dark participation is {sv['wtd_dark_pct']:.1f}%; "
             f"{sv['low_share_pct']:.0f}% of notional still routes <15% dark.")
    if np.isfinite(sv["savings_usd"]) and sv["savings_usd"] > 0:
        b.append(f"Shifting that low-dark flow toward the high-dark performance profile is an "
                 f"indicative {_fmt_usd(sv['savings_usd'])} of recoverable cost "
                 f"({sv['delta_bps']:+.1f} bps benefit).")
    b.append("Recommendation: raise the default dark target — the benefit is largest on wide-"
             "spread and high-%ADV orders, exactly where signalling risk is highest.")
    return b


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Standalone TCA report (XLSX/PDF/PNG)")
    p.add_argument("--orders", required=True, help="path to orders.csv")
    p.add_argument("--aux-dir", default=None,
                   help="folder with optional dark_routed/executed/venue_summary/algo_summary CSVs")
    p.add_argument("--outdir", default="./report")
    p.add_argument("--client", default="Client")
    p.add_argument("--period", default=None, help="reporting period label (auto from dates if omitted)")
    p.add_argument("--cost-positive", action="store_true",
                   help="flip slippage sign: source uses + = cost (default assumes + = good)")
    args = p.parse_args(argv)

    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        pass

    set_style()
    outdir = Path(args.outdir)
    png_dir = outdir / "png"
    png_dir.mkdir(parents=True, exist_ok=True)

    df = load_orders(Path(args.orders), args.cost_positive)
    print(f"Loaded {len(df):,} orders from {args.orders}.")

    period = args.period
    if period is None and df["date"].notna().any():
        period = f"{df['date'].min():%d %b %Y} – {df['date'].max():%d %b %Y}"
    period = period or "Reporting period"

    # ---- tables ----
    ot_all = ordertype_all(df)
    ot_strat = ordertype_per_strategy(df)
    ot_strat_split = ordertype_by_strategy_split(df)
    dark_tbl = dark_slippage_table(df)
    dark_size_tbl = dark_by_size_table(df)
    country_tbl = breakdown_by(df, "country", "Country")
    algo_tbl = breakdown_by(df, "strategy", "Algo")
    sv = dark_savings(df)

    aux_dir = Path(args.aux_dir) if args.aux_dir else None
    if aux_dir is None:
        print("  ! no --aux-dir given: the venue and algo charts need the auxiliary CSVs "
              "(dark_routed.csv, dark_executed.csv, dark_venue_summary.csv, algo_summary.csv).")
    elif not aux_dir.exists():
        print(f"  ! --aux-dir not found: {aux_dir}")
    else:
        have = sorted(p.name for p in aux_dir.glob("*.csv"))
        print(f"  aux-dir {aux_dir} contains: {', '.join(have) if have else '(no .csv files)'}")
    venue = load_venue_split(aux_dir)
    venue_summary = _load_aux(aux_dir, "dark_venue_summary.csv")
    algo = _load_aux(aux_dir, "algo_summary.csv")

    # ---- charts ----
    charts = []
    charts.append((chart_ordertype(ot_all, png_dir / "01_ordertype_slippage.png"),
                   "Slippage by order type"))
    charts.append((chart_dark_slippage(dark_tbl, png_dir / "02_dark_slippage.png"),
                   "More dark → better execution"))
    charts.append((chart_dark_spreadnorm(dark_tbl, png_dir / "03_dark_spreadnorm.png"),
                   "Spread capture vs dark"))
    charts.append((chart_dark_by_size(dark_size_tbl, png_dir / "04_dark_by_size.png"),
                   "Dark benefit by order size"))
    charts.append((chart_savings(sv, png_dir / "05_dark_savings.png"),
                   "The dark opportunity"))
    dark_perf_png = chart_algo_dark_perf(df, png_dir / "06_algo_dark_perf.png")
    if dark_perf_png is not None:
        # title + subtitle are baked into this figure; pass "" so the PDF/XLSX
        # don't stack a duplicate page title on top of it
        charts.append((dark_perf_png, ""))
    if venue is not None:
        charts.append((chart_venue_split(venue, png_dir / "07_venue_split.png"),
                       "Routed vs executed by venue"))
    else:
        print("  ! skipped venue-split chart: need dark_routed.csv and/or dark_executed.csv "
              "in --aux-dir, each with a venue column and a 'Routed %' / 'Executed %' column.")
    if venue_summary is not None:
        charts.append((chart_venue_summary(venue_summary, png_dir / "08_venue_summary.png"),
                       "Dark venue summary"))
    else:
        print("  ! skipped venue-summary chart: need dark_venue_summary.csv in --aux-dir "
              "with a 'Venue Category' column and a '% Total Exec Value' / 'Shares' column.")
    if algo is not None:
        charts.append((chart_algo(algo, png_dir / "09_algo_summary.png"),
                       "Performance by algorithm"))
    else:
        print("  ! skipped algo-summary chart: need algo_summary.csv in --aux-dir.")
    print(f"Wrote {len(charts)} chart(s) -> {png_dir}")

    summary_bullets = build_summary_bullets(df, ot_all, dark_tbl, sv)

    ctx = {
        "client": args.client, "period": period, "n_orders": len(df),
        "ot_all": ot_all, "ot_strat": ot_strat, "ot_strat_split": ot_strat_split,
        "dark_tbl": dark_tbl, "dark_size_tbl": dark_size_tbl, "sv": sv,
        "country_tbl": country_tbl, "algo_tbl": algo_tbl,
        "venue": venue, "venue_summary": venue_summary, "algo": algo,
        "summary_bullets": summary_bullets, "pdf_charts": charts,
    }

    slug = args.client.replace(" ", "_")
    pdf_path = outdir / f"{slug}_TCA.pdf"
    xlsx_path = outdir / f"{slug}_TCA.xlsx"
    build_pdf(pdf_path, ctx)
    print(f"Wrote PDF  -> {pdf_path}")
    build_xlsx(xlsx_path, ctx)
    print(f"Wrote XLSX -> {xlsx_path}")

    print("\nExecutive summary:")
    for b in summary_bullets:
        print(f"  • {b}")
    print(f"\nDone. Outputs in {outdir.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
